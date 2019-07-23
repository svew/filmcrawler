
from imdb import IMDb
from imdb.Person import Person
from imdb.Movie import Movie
import pymysql
import os
import openpyxl as xl

"""
Rows:
Year		MovieID	Title		Rememberance		Rating

"""

# Excel spreadsheet name
EXCEL_SPREADSHEET_NAME = 'movies.xlsx'
USER_SPREADSHEET_NAME = 'user_movies.xlsx'

# Movie Statuses
NOT_SEEN = 0
SEEN = 1
WANT_TO_SEE = 2


def search_movie(name, year):
	results = ia.search_movie(str(name))
	for movie in results:
		if str(movie['year']) == str(year):
			return movie
	return None
	
def store_user_input():
	wb_user = xl.load_workbook(USER_SPREADSHEET_NAME)
	ws_user = wb_user.active
	row = 2
	
	try:
		while True:
			year = ws_user.cell(row=row, column=1).value
			title = ws_user.cell(row=row, column=2).value
			
			retention = ws_user.cell(row=row, column=3).value
			retention = 'NULL' if retention is None else retention
			rating = ws_user.cell(row=row, column=4).value
			rating = 'NULL' if rating is None else rating
			
			row += 1
			
			if not year:
				break
			movie = search_movie(title, year)
			if not movie:
				print('Could not find movie "{}"'.format(title))
				continue
			result = db_cursor.execute('SELECT * FROM user_movies WHERE movie_id={}'.format(movie.movieID))
			if result > 0:
				print('Already found entry for "{}"'.format(title))
				continue
			values = '({}, {}, {})'.format(movie.movieID, retention, rating)
			db_cursor.execute(DB_INSERT.format('user_movies',USER_MOVIES_FIELDS,values))
			print('Stored {}'.format(title))
		
		db.commit()
	except:
		db.rollback()
		
def add_person(person):
	id = person.personID
	result = db_cursor.execute('SELECT * FROM person_general WHERE person_id={}'.format(id))
	if result > 0:
		print('Already found entry for {}'.format(person['name']))
		return
	values = '({}, {}, {})'.format(id, person['name'], 'NULL')
	db_cursor.execute(DB_INSERT.format('person_general', PERSON_GENERAL_FIELDS, values)

def setup_excel_sheets():
	global wb
	global ws_input, ws_movies, ws_people

	if not os.path.isfile(EXCEL_SPREADSHEET_NAME):
		wb = xl.Workbook()
		ws_input = wb.active
		ws_input.title = 'User Input'
		ws_movies = wb.create_sheet('Movie Data')
		ws_people = wb.create_sheet('Person Data')
	else:
		wb = xl.load_workbook(EXCEL_SPREADSHEET_NAME)
		ws_input = wb['User Input']
		ws_movies = wb['Movie Data']
		ws_people = wb['Person Data']

	ws_input['A1'] = 'MovieID'
	ws_input['B1'] = 'Year'
	ws_input['C1'] = 'Movie Title'

ia = imdb.IMDb()
setup_excel_sheets()
store_user_input()
wb.save(EXCEL_SPREADSHEET_NAME)
db.close()

